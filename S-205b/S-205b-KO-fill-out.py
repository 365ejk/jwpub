import sys
from openpyxl import load_workbook
from datetime import datetime

# load the workbook and select the sheet
input_data = sys.argv[1]  # Replace with your Excel file path
_output_template = './template/S-205b-KO-template.xlsx'
_output_path = './output/'

# read input excel form
workbook1 = load_workbook(filename=input_data)
sheet1 = workbook1.active

# parse data and update a template
for _row in sheet1.iter_rows(min_row=2):
    # read output excel template
    workbook2 = load_workbook(filename=_output_template)
    sheet2 = workbook2.active
    
    _rdate = datetime.strptime(str(_row[1].value), '%Y-%m-%d %H:%M:%S')
    _rdate = _rdate.strftime('%m/%d/%Y')
    _dmonth = _row[5].value.split(';')
    _dmonth = [_r.replace('월','') for _r in _dmonth if _dmonth]
    _dmonth = str(','.join(_dmonth)).rstrip(',')
    _rname = _row[6].value.strip().upper()

    # '_dmonth' is desired Month(s) to participate 
    _continue = _dmonth.find('다시 통치할 때까지 계속 보조 파이오니아로서 봉사하기를 원한다면 여기에 표시하십시오.')
    if _continue == -1:
        sheet2['A5'] = _dmonth
    else:
        sheet2['A5'] = '1,2,3,4,5,6,7,8,9,10,11,12'
        sheet2['A7'] = 'X'

    # '_rdate is requested date for 'S-205b' template
    sheet2['C12'] = _rdate

    # '_name' is requestor's full name
    sheet2['H15'] = _rname

    print("{0},{1},{2}".format(_rname,_dmonth,_rdate))
    
    # Step 3: Save the workbook
    _output_file = "{0}/{1}-{2}".format(_output_path,_rname,'S-205b-KO.xlsx')
    workbook2.save(filename=_output_file)
    workbook2.close()

